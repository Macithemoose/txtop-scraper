from tqdm import tqdm
import csv
from datetime import datetime
import re
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import seleniumwire.undetected_chromedriver as uc
from contextlib import contextmanager
from bs4 import BeautifulSoup
import time
import random
import os
import platform

# Place to store current headers
headers_file = 'headers_input.txt'

# Plane class; attributes are all the things I want to scrape.
class Plane:
    def __init__(self, price=None, manufacturer=None, model=None, year=None, flight_hours=None, engine_time=None, detail_url=None, tail_number=None, avionics=None):
        self.price = price
        self.manufacturer = manufacturer
        self.model = model
        self.year = year
        self.flight_hours = flight_hours # airframe total time
        self.engine_time = engine_time # engine total time
        self.detail_url = detail_url
        self.tail_number = tail_number
        self.avionics = avionics

    def __repr__(self):
        return (f"Plane(manufacturer={self.manufacturer}, model={self.model}, year={self.year}, "
                f"price={self.price}, detail_url={self.detail_url}, engine_time = {self.engine_time}," 
                f"tail_number = {self.tail_number}, avionics = {self.avionics})")

def extract_planes_from_listings(listings):
    planes = []
    for listing in listings:
        price = listing.get("RetailPrice", "N/A")
        manufacturer = listing.get("ManufacturerName", "N/A")
        model = listing.get("Model", "N/A")
        year = None
        flight_hours = None
        engine_time = None
        detail_url = "https://controller.com" + listing.get("DetailUrl", "N/A") + " "
        tail_number = None
        avionics = None

        # Get the year:
        title = listing.get("ListingTitle", "")
        year_match = re.search(r"\b(19|20)\d{2}\b", title)
        if year_match:
            year = year_match.group()

        # Extract year, airframe time, engine time, and tail number from specs
        specs = listing.get("Specs", [])
        for spec in specs:
            if spec["Key"] == "Total Time":
                flight_hours = spec["Value"]
                flight_hours = flight_hours.replace(",", "")
                flight_hours = float(flight_hours) if isinstance(flight_hours, str) else flight_hours
            elif spec["Key"] == "Engine 1 Time":
                engine_time = spec["Value"]
            elif spec["Key"] == "Registration #":
                tail_number = spec["Value"]

        # Create a Plane object
        plane = Plane(
            price=price,
            manufacturer=manufacturer,
            model=model,
            year=year,
            flight_hours=flight_hours,
            engine_time=engine_time,
            detail_url=detail_url,
            tail_number=tail_number,
            avionics=avionics
        )
        planes.append(plane)
    
    return planes

# Keep letters, numbers, whitespace (\s), and newlines
def clean_text(text):
    if not text:
        return ""
    return re.sub(r"[^\w\s\r\n:/]", "", text)

# Remove non-printable characters
def clean_avionics(text):
    if not text:
        return ""

    text = re.sub(r"[^\x20-\x7E\r\n]", "", text)
    
    lines = text.splitlines()
    cleaned_lines = [line.strip() for line in lines if line.strip()]
    return "\r\n".join(cleaned_lines)

# Exports all the planes to an xslx file
def export_planes_to_xlsx(planes, output_file):
    headers = [
        "price", "manufacturer", "model", "year", "airframe_total_time", 
        "engine_total_time","detail_url","tail number", "avionics"
    ]

    wb = Workbook()
    ws = wb.active
    ws.append(headers)

    col_widths = [30, 15, 20, 15, 10, 15, 15, 50, 15, 50]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for plane in planes:
        avionics_text = plane.avionics.replace('\n', '\r\n') if plane.avionics else ""
        avionics_text = clean_text(avionics_text)
        avionics_text = clean_avionics(avionics_text)
        plane.avionics = avionics_text

        ws.append([
            plane.price,
            plane.manufacturer,
            plane.model,
            plane.year,
            plane.flight_hours,
            plane.engine_time,
            plane.detail_url,
            plane.tail_number,
            plane.avionics
        ])

        for row in ws.iter_rows(min_row=2, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
    wb.save(output_file)
    print(f"Data exported to {output_file}")


# Write listing into CSV file:
def export_planes_to_csv(planes, output_file):
    headers = [
        "price", "manufacturer", "model", "year", "airframe_total_time", 
        "engine_total_time","detail_url","tail number", "avionics"
    ]
    
    with open(output_file, mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        for plane in planes:
            avionics_text = plane.avionics.replace('\n', '\r\n') if plane.avionics else ""
            avionics_text = clean_text(avionics_text)
            avionics_text = clean_avionics(avionics_text)

            plane.avionics = avionics_text

            writer.writerow([
                plane.price,
                plane.manufacturer,
                plane.model,
                plane.year,
                plane.flight_hours,
                plane.engine_time,
                plane.detail_url,
                plane.tail_number,
                plane.avionics
            ])
    print(f"Data exported to {output_file}")

# Converts the headers extracted from the website into a dictionary that can then be fed to the driver to implement the headers
def convert_headers_to_dict(headers_file):
    with open(headers_file, "r") as file:
        curl_command = file.read()

    headers_pattern = r"-H '([^:]+): (.*?)'"
    headers_matches = re.findall(headers_pattern, curl_command)

    headers = {}
    for key, value in headers_matches:
        if key.lower() == "cookie":
            headers[key.lower()] = " ".join(value.splitlines())
        else:
            headers[key.lower()] = value
    # print(headers)
    return headers

# Loads the last time we updated headers to check whether we need to update them again
def load_last_update_time():
    try:
        with open("last_update.txt", "r") as f:
            timestamp_str = f.read().strip()
            if not timestamp_str:
                return None
            return datetime.fromisoformat(timestamp_str)
    except FileNotFoundError:
        return None

# Updates last_update.txt with the latest headers scrape time
def save_last_update_time(dt):
    with open("last_update.txt", "w") as f:
        f.write(dt.isoformat())
        
# Gets the avionics for a specific plane using its detail_url
def extract_avionics(planes):
    driver = setup_driver()
    driver.request_interceptor = interceptor

    try:
        for plane in tqdm(planes, desc="Scraped Avionics", bar_format="{desc}: {n_fmt}/{total_fmt} •{bar:30}•"):
            time.sleep(random.uniform(1, 5))
            url = plane.detail_url
            # clear other reqests before searching:
            if hasattr(driver, "requests"):
                driver.requests.clear()

            driver.get(url)

            # page_html = driver.page_source.lower()
            # if any(term in page_html for term in ["Pardon Our Interruption"]):
            #     print(f"[ERROR] Bot detection triggered for {url}. Try switching the VPN or running headers again. Stopping scraper.")
            #     plane.avionics = "Bot detection page encountered"
            #     break

            try:
                wait = WebDriverWait(driver, 30)
                avionics_heading = wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//h3[contains(@class,'detail__specs-heading') and contains(., 'Avionics')]")
                    )
                )
            except:
                # If we never see the heading, skip
                plane.avionics = "No avionics info found"
                continue

            soup = BeautifulSoup(driver.page_source, 'html.parser')

            h3_heading = soup.find("h3", class_="detail__specs-heading", string=re.compile(r"Avionics", re.IGNORECASE))

            if h3_heading:
                wrapper = h3_heading.find_next_sibling("div", class_="detail__specs-wrapper")
                if wrapper:
                    labels = wrapper.find_all("div", class_="detail__specs-label")
                    values = wrapper.find_all("div", class_="detail__specs-value")
                    if labels and values and len(labels) == len(values):
                        pairs = []
                        for l, v in zip(labels, values):
                            label = l.get_text()
                            value = v.get_text()
                            pair = f"{label}: {value}"
                            pairs.append(pair)
                        plane.avionics = "\n".join(pairs)
                        continue
                    else:
                        plane.avionics = "Avionics section found, but label/value mismatch"
                        continue
                else:
                    plane.avionics = "Avionics heading found, but wrapper not found"
                    continue
            else:
                plane.avionics = "Could not find an Avionics heading in the HTML"

    except Exception as e:
        print(f"Ran into error: {e}")
    finally:
        driver.quit()
        del driver

# If the user inputs years, this reduces the list of planes to only planes in that range of user-inputted years.
def filter_year(planes, years):
    correct_planes = []
    for plane in planes:
        if plane.year in years:
            correct_planes.append(plane)
    return correct_planes


# Sets up the web driver with the right options and headers
# def setup_driver():
    options = uc.ChromeOptions()
    options.add_argument('--allow-insecure-localhost')
    options.add_argument("--enable-javascript")
    options.add_argument('--allow-third-party-cookies')

    # If running in Docker (Linux), the Dockerfile sets these env vars:
    chrome_bin = os.environ.get("CHROME_BIN")
    selenium_flags = os.environ.get("SELENIUM_FLAGS", "")

    # Apply any headless/container flags from env (space-separated)
    for flag in selenium_flags.split():
        if flag.strip():
            options.add_argument(flag.strip())

    caps = options.to_capabilities()
    caps["acceptInsecureCerts"] = True

    # Choose executable path depending on environment
    exec_path = None
    if chrome_bin and platform.system().lower() == "linux":
        # Docker/CI: trust env path to Chromium
        exec_path = chrome_bin
        driver = uc.Chrome(browser_executable_path=exec_path, options=options, desired_capabilities=caps)
        print('Used Linux chrome')
    else:
        # Local Windows dev fallback (your original behavior)
        print("Used original Windows chrome")
        exec_path = "chrome-win64/chrome.exe"
        driver = uc.Chrome(browser_executable_path=exec_path, version_main=134, options=options, desired_capabilities=caps)

    return driver

def setup_driver():
    options = uc.ChromeOptions()  
    # adding option to get past insecure network error:
    options.add_argument('--allow-insecure-localhost') # differ on driver version. can ignore. 
    caps = options.to_capabilities()
    caps["acceptInsecureCerts"] = True
    # adding option to deal with bot detection error: javascript not enabled
    options.add_argument("--enable-javascript")
    # adding option to enable cookies:
    options.add_argument('--allow-third-party-cookies')
    #options.add_argument('--headless') # this actually activates bot detection

    ## Create Chrome Driver
    driver = uc.Chrome(
        browser_executable_path="chrome-win64/chrome.exe", 
        version_main = 134, 
        options=options, 
        desired_capabilities=caps)
    return driver

# Create a request interceptor to modify headers to escape bot detection
def interceptor(request):
    del request.headers['accept-language']
    del request.headers['accept-encoding']
    request.headers['accept'] =  "application/json, text/plain, */*"
    request.headers['accept-language'] = 'en-US,en;q=0.9'
    request.headers['accept-encoding'] = 'gzip, deflate, br, zstd'
    request.headers['content-type'] = 'application/json'
    request.headers['priority'] = 'u=1, i'
    request.headers['referer'] = 'https://www.controller.com/'
    #request.headers['sec-ch-ua'] = '"Google Chrome";v="134", "Chromium";v="134", "Not_A Brand";v="24"'
    request.headers['sec-ch-ua-mobile'] = '?0'
    request.headers['sec-ch-ua-platform'] = 'Windows' if platform.system() == 'Windows' else 'Linux'
    request.headers['sec-fetch-dest'] = 'empty'
    request.headers['sec-fetch-mode'] = 'cors'
    request.headers['sec-fetch-site'] = 'same-origin'

# Takes the scraped cookies and updates the headers file with them
def write_headers_with_updated_cookies(cookies, cookies_to_replace, filename):
    with open(filename, "r") as f:
        headers_content = f.read()

    # 2. Regex to match the -H 'cookie: ...' line
    pattern = r"(-H\s*'cookie:\s*)([^']*)(')"
    match = re.search(pattern, headers_content)
    if not match:
        print("No cookie header found. Writing file back unchanged.")
        return

    prefix = match.group(1)  # e.g. "-H 'cookie: "
    cookie_str = match.group(2)  # the actual cookie string
    suffix = match.group(3)  # trailing single quote

    # 3. Parse existing cookies into a dict
    cookie_map = {}
    for part in cookie_str.split(';'):
        part = part.strip()
        if '=' in part:
            k, v = part.split('=', 1)
            cookie_map[k.strip()] = v.strip()

    # 4. Update only the cookies we have in our list
    for c in cookies:
        name, value = c["name"], c["value"]
        if name in cookie_map:
            cookie_map[name] = value

    # 5. Rebuild the cookie string
    new_cookie_str = "; ".join(f"{k}={v}" for k, v in cookie_map.items())

    # 6. Construct the updated line
    updated_cookie_line = prefix + new_cookie_str + suffix

    # 7. Splice the updated cookie line back into the file contents
    updated_content = (
        headers_content[:match.start()] +
        updated_cookie_line +
        headers_content[match.end():]
    )

    # 8. Overwrite the file with updated content
    with open(filename, "w") as out:
        out.write(updated_content)

    # print(f"Successfully updated cookies in {filename}.")

# Updates the x-xsrf-token, once found
def update_x_security(token, filename):
    with open(filename, "r") as f:
        content = f.read()
    pattern = r"(-H\s*'x-xsrf-token:\s*)([^']*)(')"

    def replacer(match):
        prefix = match.group(1)  # e.g. -H 'x-xsrf-token:
        # old_value = match.group(2)  # The current token (not used)
        suffix = match.group(3)  # The trailing single quote
        return f"{prefix}{token}{suffix}"  # Insert our new token

    updated_content = re.sub(pattern, replacer, content)

    with open(filename, "w") as f:
            f.write(updated_content)


# Waits for the x-xsrf-token to appear in driver.requests, and returns it (if found)
def wait_for_token_passive(driver, timeout=30):
    end_time = time.time() + timeout
    while time.time() < end_time:
        # Look at recent requests
        for req in reversed(driver.requests):
            if req.headers.get('x-xsrf-token'):
                return req.headers['x-xsrf-token']
        time.sleep(0.2)  # Gentle poll to avoid CPU spike

    print("⚠️ Timeout: x-xsrf-token not found.")
    return None

# Functions for updating the x xsrf token if not found immediately:
def xsrf_from_cookies(cookies):
    # Look for common anti-CSRF cookie names
    for c in cookies:
        n = (c.get("name") or "").lower()
        if n in ("xsrf-token", "__xsrf-token", "__requestverificationtoken"):
            return c.get("value")
    return None

headers_path = 'headers_input.txt'
# Updates the headers, using the functions above.
def update_headers():
    driver = setup_driver()
    driver.request_interceptor = interceptor
    try:
        driver.get('https://www.controller.com/listings/search?page=3')
        # driver.requests.clear()
        x_xsrf_token = wait_for_token_passive(driver, timeout=30)

        cookies = driver.get_cookies()

        if not x_xsrf_token:
            cookie_token = xsrf_from_cookies(cookies)
            if cookie_token:
                x_xsrf_token = cookie_token
                print("Using XSRF token from cookies.")
            else:
                print("Still couldn't find x xsrf token.")
            print("Still couldn't find x xsrf token.")
        else:
            print(f"Found it!")

        cookies_to_replace = {
            "_hjSessionUser_1143839",
            "ASP.NET_SessionId",
            "__RequestVerificationToken",
            "__XSRF-TOKEN",
            "Tracking",
            "reese84",
            "UserID",
            "VisitorID"
        }

        write_headers_with_updated_cookies(cookies, cookies_to_replace, headers_path)
        update_x_security(x_xsrf_token, headers_path)
        print("Headers updated")

    except Exception as e:
        print(f"Error during running driver: {e}")

    finally:
        driver.quit()
        del driver

def noop(*args, **kwargs):
    pass

uc.Chrome.__del__ = noop
